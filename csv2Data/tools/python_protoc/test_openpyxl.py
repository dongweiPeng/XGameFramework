#coding: utf-8
# 依赖openpyxl库：http://openpyxl.readthedocs.org/en/latest/

from openpyxl import Workbook
from openpyxl import load_workbook
import csv
import os
import sys
import codecs

def xlsx2csv(filename, file_dir, export_csv_path):
    file_path = file_dir + filename
    try:
        xlsx_file_reader = load_workbook(filename=file_path)
        for sheet in xlsx_file_reader.get_sheet_names():
            # 每个sheet输出到一个csv文件中，文件名用xlsx文件名和sheet名用'_'连接
            csv_filename = sheet + '.csv'
            xlsx_data = []
            sheet_ranges = xlsx_file_reader[sheet]
            if sheet_ranges.max_row < 3:
        #        print '%s is empty!' % (csv_filename)
                continue

            for row in sheet_ranges.rows:
                row_container = []
                for cell in row:
                    cell_val = cell.value
                    #将空值转成空字符串
                    if cell.value == None:
                        cell_val = ''
                    elif type(cell.value) == unicode:
                        cell_val = cell.value.encode('utf-8')
                    else:
                        cell_val = str(cell.value)
                    row_container.append(cell_val)
                xlsx_data.append(row_container)
            #存入csv文件
            csv_file = file(export_csv_path + csv_filename, 'wb')
            csv_file.write(codecs.BOM_UTF8)
            csv_file_writer = csv.writer(csv_file)
            for row_data in xlsx_data:
                csv_file_writer.writerow(row_data)
            csv_file.close()
            print 'export csv file: %s succ~' % (csv_filename)

    except Exception as e:
        print 'export excel file %s to csv failed!' % (filename)
        print(e)
